import openpyxl
import dropbox
from twilio.rest import Client
from datetime import datetime, timedelta
import io


# Inicialização da lista global de dados
lista_clientes = []
lista_nmr_clientes = []
lista_valores = []
lista_porcentagem = []
lista_pagamento = []
lista_parcelas = []
lista_valor_parcelas = []
lista_valor_total = []
lista_data_cadastro = []

# Função para coletar os dados do cliente
def coletar_dados_cliente():
    cliente = input("Digite o nome do cliente: ")
    lista_clientes.append(cliente)

    nmr_cliente = input("Digite o número do cliente (+55 + DDD + Cel): ")
    lista_nmr_clientes.append(nmr_cliente)

    valor = int(input("Digite o valor emprestado ao cliente: "))
    lista_valores.append(valor)

    porcentagem = int(input("Digite o valor da porcentagem de juros: "))
    lista_porcentagem.append(porcentagem)

    intervalo_pagamentos = int(input("Digite a cada quanto tempo o cliente irá pagar (em dias): "))
    lista_pagamento.append(intervalo_pagamentos)

    parcelas = int(input("Digite o número de parcelas: "))
    lista_parcelas.append(parcelas)

    valor_total = valor
    valor_parcela = (valor_total / parcelas) * (1 + porcentagem / 100)
    formatted_valor_parcela = round(valor_parcela, 2)
    lista_valor_parcelas.append(formatted_valor_parcela)

    total_a_pagar = valor_total * (1 + porcentagem / 100)
    lista_valor_total.append(total_a_pagar)

    data_hoje = datetime.now()
    data_formatada = data_hoje.strftime("%d/%m/%Y")  # Modificado para incluir o ano com quatro dígitos
    lista_data_cadastro.append(data_formatada)

    # Retorne os valores relevantes
    return cliente, nmr_cliente, valor, porcentagem, intervalo_pagamentos, parcelas, formatted_valor_parcela, total_a_pagar, data_formatada

# Função para fazer upload para o Dropbox
def adicionar_dados_planilha_local():
    # Nome do arquivo da planilha na pasta do projeto
    nome_arquivo = 'clientes.xlsx'

    # Tente carregar a planilha existente
    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        # Se a planilha não existir, crie uma nova
        workbook = openpyxl.Workbook()

    # Obtenha a folha ativa
    sheet = workbook.active

    # Encontre a próxima linha vazia na planilha
    proxima_linha = sheet.max_row + 1

    # Adicione os dados dos clientes à planilha a partir da próxima linha vazia
    for i in range(len(lista_clientes)):
        sheet.cell(row=proxima_linha + i, column=1, value=lista_clientes[i])
        sheet.cell(row=proxima_linha + i, column=2, value=lista_nmr_clientes[i])
        sheet.cell(row=proxima_linha + i, column=3, value=lista_valores[i])
        sheet.cell(row=proxima_linha + i, column=4, value=lista_porcentagem[i])
        sheet.cell(row=proxima_linha + i, column=5, value=lista_pagamento[i])
        sheet.cell(row=proxima_linha + i, column=6, value=lista_parcelas[i])
        sheet.cell(row=proxima_linha + i, column=7, value=lista_valor_parcelas[i])
        sheet.cell(row=proxima_linha + i, column=8, value=lista_valor_total[i])
        sheet.cell(row=proxima_linha + i, column=9, value=lista_data_cadastro[i])

    # Salve a planilha atualizada na pasta do projeto
    workbook.save(nome_arquivo)


def calculo_parcelas():
    # Nome do arquivo da planilha na pasta do projeto
    nome_arquivo = 'clientes.xlsx'

    # Tente carregar a planilha existente
    try:
        workbook = openpyxl.load_workbook(nome_arquivo)
    except FileNotFoundError:
        print("A planilha 'clientes.xlsx' não foi encontrada.")
        return []

    # Obtenha a folha ativa
    sheet = workbook.active

    # Data de hoje
    data_atual = datetime.now().date()

    # Lista para armazenar as informações de envio de SMS
    infos_envio_sms = []

    # Percorra as linhas da planilha
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_cadastro_str = row[8]  # Coluna 9 contém a data de cadastro

        # Verifique se o valor da célula é None ou vazio
        if data_cadastro_str is None or data_cadastro_str == "":
            print("Data de cadastro ausente ou vazia.")
            continue

        # Verifique se o valor da célula é um objeto datetime
        if isinstance(data_cadastro_str, datetime):
            data_cadastro = data_cadastro_str.date()  # Converta para data se for datetime
        else:
            data_cadastro_str = str(data_cadastro_str)
            try:
                data_cadastro = datetime.strptime(data_cadastro_str, "%d/%m/%Y").date()  # Sempre com AAAA
            except ValueError:
                print(f"Erro na data de cadastro: '{data_cadastro_str}'. Verifique o formato (DD/MM/AAAA).")
                continue

        intervalo_pagamento = row[4]  # Coluna 5 contém o intervalo de pagamento em dias
        nome_cliente = row[0]  # Coluna 1 contém o nome do cliente
        telefone_cliente = row[1]  # Coluna 2 contém o telefone do cliente
        valor_parcela = row[6]  # Coluna 7 contém o valor da parcela

        # Calcule a diferença de meses desde a data de cadastro até a data atual
        meses_desde_cadastro = (data_atual.year - data_cadastro.year) * 12 + (data_atual.month - data_cadastro.month)

        # Verifique se a parcela deve ser paga hoje (mês seguinte ao cadastro)
        if meses_desde_cadastro % intervalo_pagamento == 0:
            numero_parcela = meses_desde_cadastro // intervalo_pagamento + 1
            infos_envio_sms.append((nome_cliente, telefone_cliente, valor_parcela, numero_parcela))

    return infos_envio_sms

def sms_planilha_info():
    # Configurar suas credenciais do Twilio
    account_sid = 'ACae6b3430341fde63009bb4ccb9881310'
    auth_token = '0c517bd637320f5e88532f7ad523f3b9'
    twilio_phone_number = '+14783304454'
    client_twilio = Client(account_sid, auth_token)

    # Configurar suas credenciais do Dropbox
    access_token_dropbox = 'sl.Bm8WoIrwop3ju77mTCaTV-4zjZgSACEek2HAVGO_LkKN7mQthp-iho3ERaTxdgJlqu0XZewqQ8sCfQvatsbbFDVuo0hida5T-YTYnJmNCH6NZvvZqr9OQuQ-3xkDio-SbMVAwMCf1HWFcJfXehjBv5I'
    dbx = dropbox.Dropbox(access_token_dropbox)

    # Nome do arquivo da planilha no Dropbox
    nome_arquivo_dropbox = '/clientes.xlsx'

    # Baixar a planilha do Dropbox e lê-la com o openpyxl
    _, response = dbx.files_download(nome_arquivo_dropbox)
    conteudo_planilha = response.content

    # Abrir a planilha usando o openpyxl
    planilha = openpyxl.load_workbook(io.BytesIO(conteudo_planilha), data_only=True)
    sheet = planilha.active

    # Inicializar a lista de dados dos clientes
    dados_clientes = []

    # Ler os dados da planilha a partir da segunda linha (pulando o cabeçalho)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados_clientes.append(row)

    # Data atual
    data_hoje = datetime.now()

    # Mostrar as entradas dos clientes
    for idx, row in enumerate(dados_clientes, start=2):
        nome_cliente, numero_celular, _, _, _, parcela, valor_parcela, _, data_vencimento = row
        data_vencimento = datetime.strptime(data_vencimento, '%d/%m/%y')  # Corrigir a formatação da data
        print(f"{idx - 1}. Nome: {nome_cliente}, Celular: {numero_celular}, Parcela: {parcela}, Valor: R${valor_parcela:.2f}, Vencimento: {data_vencimento.strftime('%d/%m/%Y')}")

    # Solicitar a seleção do cliente
    while True:
        try:
            indice_cliente = int(input("Digite o índice do cliente para enviar a mensagem (0 para sair): "))
            if 1 <= indice_cliente <= len(dados_clientes):
                break
            else:
                print("Índice inválido. Por favor, escolha um índice válido.")
        except ValueError:
            print("Entrada inválida. Digite um número válido.")

    if indice_cliente == 0:
        print("Saindo do programa.")
        return

    # Cliente selecionado
    nome_cliente, numero_celular, _, _, _, parcela, valor_parcela, _, data_vencimento = dados_clientes[indice_cliente - 1]
    data_vencimento = datetime.strptime(data_vencimento, '%d/%m/%y')  # Corrigir a formatação da data

    # Calcular as datas de envio de SMS apenas para a próxima parcela a vencer
    datas_envio_sms = calculo_parcelas(data_vencimento, int(parcela), int(parcela))
    data_proxima_parcela = datas_envio_sms[0]

    # Enviar a mensagem de SMS para a próxima parcela a vencer
    data_envioSMS_formatada = data_proxima_parcela.strftime("%d/%m/%Y")
    mensagem = f"Olá {nome_cliente}, você está recebendo um aviso em relação à próxima parcela no valor de R${valor_parcela:.2f} com vencimento em {data_envioSMS_formatada}. Providencie o pagamento."

    # Enviar a mensagem de SMS
    message = client_twilio.messages.create(
        body=mensagem,
        from_=twilio_phone_number,
        to='+' + numero_celular
    )

    print(f"Mensagem enviada para {nome_cliente}: {message.sid}")


